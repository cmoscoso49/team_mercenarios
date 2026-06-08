import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\cmoscoso\OneDrive - INACAP\Descargas\2025\Mercenarios\datos team Actualizada 2022-2025 REVISADO POR ESTEBAN TL (4).xlsx", read_only=True, data_only=True)
print("Hojas:", wb.sheetnames)
print()

for hoja in ["MENSUALIDADES 2024", "MENSUALIDADES 2025"]:
    if hoja not in wb.sheetnames:
        print(f"HOJA NO ENCONTRADA: {hoja}")
        continue
    ws = wb[hoja]
    print(f"=== {hoja} ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        vals = [str(v)[:18] if v is not None else "" for v in row[:24]]
        print(f"  Fila {i}: {vals}")
    print()
wb.close()