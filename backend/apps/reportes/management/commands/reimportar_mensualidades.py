"""
Reimporta mensualidades 2024 y 2025 desde el Excel usando matching mejorado.
- Normaliza acentos y case antes de comparar nicks
- Fallback a nombre exacto si el nick no matchea
- Idempotente: usa update_or_create, no duplica datos
- No toca otros años ni integrantes

Uso:
    python manage.py reimportar_mensualidades
    python manage.py reimportar_mensualidades --dry-run
    python manage.py reimportar_mensualidades --solo-anio 2025
"""
import unicodedata
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction


EXCEL_PATH = r"C:\Users\cmoscoso\OneDrive - INACAP\Descargas\2025\Mercenarios\datos team Actualizada 2022-2025 REVISADO POR ESTEBAN TL (4).xlsx"

MESES_MAP = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JUNIO ': 6, 'JULIO': 7,
    'AGOSTO': 8, 'AGOSTO ': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10,
    'NOBIEMBRE': 11, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
}

FILAS_IGNORAR_KEYWORDS = (
    'DEFINICION', 'ACTIVO :', 'INACTIVOS :', 'INACTIVO :', 'NOTA:', 'DEBERES',
    'PLAYERS QUE NO CANCELEN', 'MENSUALIDADES Y RESPECTIVAS',
)


def _val(v):
    if v is None:
        return ''
    return str(v).strip()


def _norm(s):
    """Normaliza: quita acentos, uppercase, strip."""
    s = unicodedata.normalize('NFD', str(s))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.upper().strip()


def _es_fila_texto(nick_raw, nombre_raw):
    """Detecta filas de reglamento/notas dentro del Excel."""
    if not nick_raw and not nombre_raw:
        return True
    texto = (nombre_raw + ' ' + nick_raw).upper()
    return any(k in texto for k in FILAS_IGNORAR_KEYWORDS)


class Command(BaseCommand):
    help = 'Reimporta mensualidades 2024 y 2025 con matching mejorado (unicode + nombre fallback).'

    def add_arguments(self, parser):
        parser.add_argument('--dry-run', action='store_true', help='Simula sin guardar')
        parser.add_argument('--solo-anio', type=int, choices=[2024, 2025], help='Solo importar un año')

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.solo_anio = options.get('solo_anio')
        self.errores = []
        self.sin_match = []

        modo = '[DRY-RUN] ' if self.dry_run else ''
        self.stdout.write(self.style.WARNING(f'{modo}Reimportando mensualidades 2024-2025...\n'))

        try:
            self.wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        except FileNotFoundError:
            self.stderr.write(f'ERROR: No se encontro el archivo:\n  {EXCEL_PATH}')
            return

        self._construir_indices()

        hojas = {}
        if not self.solo_anio or self.solo_anio == 2024:
            hojas['MENSUALIDADES 2024'] = 2024
        if not self.solo_anio or self.solo_anio == 2025:
            hojas['MENSUALIDADES 2025'] = 2025

        total_creadas = 0
        total_actualizadas = 0

        with transaction.atomic():
            for hoja_nombre, anio in hojas.items():
                creadas, actualizadas = self._procesar_hoja(hoja_nombre, anio)
                total_creadas += creadas
                total_actualizadas += actualizadas
            if self.dry_run:
                transaction.set_rollback(True)

        self.wb.close()
        self._imprimir_resumen(total_creadas, total_actualizadas)

    def _construir_indices(self):
        from apps.integrantes.models import Integrante
        integrantes = list(Integrante.objects.all())
        self.nick_index = {}
        self.nombre_index = {}
        for i in integrantes:
            if i.nick:
                self.nick_index[_norm(i.nick)] = i
            self.nombre_index[_norm(i.nombre)] = i

    def _find_integrante(self, nick_raw, nombre_raw):
        """Busca integrante por nick normalizado, luego por nombre exacto normalizado."""
        nick_n = _norm(nick_raw)
        nombre_n = _norm(nombre_raw)

        if nick_n and nick_n in self.nick_index:
            return self.nick_index[nick_n], 'nick'

        if nombre_n and nombre_n in self.nombre_index:
            return self.nombre_index[nombre_n], 'nombre'

        return None, None

    def _procesar_hoja(self, hoja_nombre, anio):
        from apps.finanzas.models import Mensualidad

        if hoja_nombre not in self.wb.sheetnames:
            self.stdout.write(f'  Hoja no encontrada: {hoja_nombre}')
            return 0, 0

        ws = self.wb[hoja_nombre]
        self.stdout.write(f'\nProcesando {hoja_nombre}...')

        # Encontrar fila de cabecera
        header_row = None
        headers = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            vals = [_val(v).upper() for v in row]
            if 'NICK' in vals or 'NOMBRE DE PLAYERS' in vals:
                header_row = i
                headers = vals
                break

        if header_row is None:
            self.stdout.write(f'  No se encontro cabecera en {hoja_nombre}')
            return 0, 0

        # Mapear columnas de meses
        mes_cols = {}
        for ci, h in enumerate(headers):
            mes = MESES_MAP.get(h.strip())
            if mes:
                mes_cols[mes] = ci

        creadas = 0
        actualizadas = 0
        matcheados = set()

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            nick_raw = _val(row[3] if len(row) > 3 else None)
            nombre_raw = _val(row[2] if len(row) > 2 else None)

            if _es_fila_texto(nick_raw, nombre_raw):
                continue

            integrante, metodo = self._find_integrante(nick_raw, nombre_raw)

            if not integrante:
                if nick_raw or nombre_raw:
                    self.sin_match.append(f"  {anio}: nick='{nick_raw}' nombre='{nombre_raw[:40]}'")
                continue

            matcheados.add(integrante.id)

            for mes, col in mes_cols.items():
                if col >= len(row):
                    continue
                cell_val = _val(row[col])

                if not cell_val:
                    estado = 'pendiente'
                elif cell_val.replace('.', '').replace(',', '').isdigit():
                    estado = 'pagada'
                elif 'GRACIA' in cell_val.upper() or 'EXENT' in cell_val.upper():
                    estado = 'exento'
                else:
                    estado = 'pendiente'

                # Ene-May 2025 fueron condonados por el team (regla de negocio)
                if anio == 2025 and mes in (1, 2, 3, 4, 5) and estado == 'pendiente':
                    estado = 'exento'

                try:
                    monto = int(float(cell_val.replace('.', '').replace(',', ''))) if estado == 'pagada' else 5000
                except (ValueError, AttributeError):
                    monto = 5000

                obj, created = Mensualidad.objects.update_or_create(
                    integrante=integrante, anio=anio, mes=mes,
                    defaults={'estado': estado, 'monto': monto},
                )
                if created:
                    creadas += 1
                else:
                    actualizadas += 1

        self.stdout.write(f'  Matcheados ({metodo or "mix"}): {len(matcheados)} integrantes')
        self.stdout.write(f'  Registros creados: {creadas} | actualizados: {actualizadas}')
        return creadas, actualizadas

    def _imprimir_resumen(self, creadas, actualizadas):
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'=== RESUMEN ==='))
        self.stdout.write(f'  Creadas: {creadas} | Actualizadas: {actualizadas}')

        if self.sin_match:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(f'=== SIN MATCH ({len(self.sin_match)}) — requieren revision manual ==='))
            for e in self.sin_match:
                self.stdout.write(e)

        if self.errores:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR(f'=== ERRORES ({len(self.errores)}) ==='))
            for e in self.errores:
                self.stdout.write(f'  {e}')

        self.stdout.write('')
        if self.dry_run:
            self.stdout.write(self.style.WARNING('[DRY-RUN] No se guardaron cambios.'))
        else:
            self.stdout.write(self.style.SUCCESS('Reimportacion completada.'))