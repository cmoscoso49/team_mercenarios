"""
Sincroniza integrantes usando la hoja "MENSUALIDADES 2025" como fuente oficial.

Estructura de la hoja (detectada del Excel real):
  Col B (índice 1): Número secuencial
  Col C (índice 2): Nombre completo
  Col D (índice 3): Nick / alias
  Col E (índice 4): Estado (ACTIVO / INACTIVO / POS NATAL)
  Datos desde fila 7 (cabecera en fila 6)

Uso:
  python manage.py sincronizar_integrantes_2025
  python manage.py sincronizar_integrantes_2025 --dry-run
"""
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

EXCEL_PATH = (
    r"C:\Users\cmoscoso\OneDrive - INACAP\Descargas\2025\Mercenarios"
    r"\datos team Actualizada 2022-2025 REVISADO POR ESTEBAN TL (4).xlsx"
)

HOJA = 'MENSUALIDADES 2025'

ESTADO_MAP = {
    'ACTIVO': 'activo',
    'INACTIVO': 'inactivo',
    'POS NATAL': 'pos_natal',
    'POST NATAL': 'pos_natal',
    'POSNATAL': 'pos_natal',
    'POSTNATAL': 'pos_natal',
    'SUSPENDIDO': 'suspendido',
    'POSTULANTE': 'postulante',
    'HONORARIO': 'honorario',
}

# Índices de columna (base 0) — confirmados desde screenshot del Excel real
IDX_NOMBRE = 2   # Col C
IDX_NICK   = 3   # Col D
IDX_ESTADO = 4   # Col E
FILA_DATOS_INICIO = 7  # Primera fila de datos (fila 7 en Excel = índice 6 en iter_rows)

STOP_KEYWORDS = ('TOTAL', 'RESUMEN', 'SUMA', 'PROMEDIO')


def _val(v):
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ('None', 'nan') else s


class Command(BaseCommand):
    help = 'Sincroniza integrantes desde la hoja MENSUALIDADES 2025 del Excel oficial.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Muestra qué se haría sin guardar cambios en la BD.'
        )

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        prefix = '[DRY-RUN] ' if self.dry_run else ''
        self.stdout.write(self.style.WARNING(
            f'{prefix}Sincronizando integrantes desde "{HOJA}"...\n'
        ))

        try:
            wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        except PermissionError:
            self.stderr.write(
                'ERROR: El archivo Excel está abierto en otro programa.\n'
                'Cierra Excel y vuelve a ejecutar el comando.'
            )
            return
        except FileNotFoundError:
            self.stderr.write(f'ERROR: Archivo no encontrado:\n  {EXCEL_PATH}')
            return

        if HOJA not in wb.sheetnames:
            self.stderr.write(
                f'ERROR: Hoja "{HOJA}" no encontrada.\n'
                f'Hojas disponibles: {wb.sheetnames}'
            )
            wb.close()
            return

        ws = wb[HOJA]
        filas = self._extraer_filas(ws)
        wb.close()

        if not filas:
            self.stderr.write('ERROR: No se encontraron filas de datos en la hoja.')
            return

        self.stdout.write(f'  Filas leídas del Excel: {len(filas)}')
        self._sincronizar(filas, prefix)

    def _extraer_filas(self, ws):
        filas = []
        for row in ws.iter_rows(min_row=FILA_DATOS_INICIO, values_only=True):
            nombre = _val(row[IDX_NOMBRE] if len(row) > IDX_NOMBRE else None)
            nick   = _val(row[IDX_NICK]   if len(row) > IDX_NICK   else None)
            estado_raw = _val(row[IDX_ESTADO] if len(row) > IDX_ESTADO else None).upper()

            if not nombre and not nick:
                continue
            # Filas sin nick Y sin estado son notas/definiciones del Excel — ignorar
            if not nick and not estado_raw:
                continue
            if any(k in nombre.upper() for k in STOP_KEYWORDS):
                break
            if nick and any(k in nick.upper() for k in STOP_KEYWORDS):
                break

            estado = ESTADO_MAP.get(estado_raw, 'activo')

            filas.append({
                'nombre': nombre[:150],
                'nick':   nick[:50],
                'estado': estado,
                'estado_raw': estado_raw,
            })
        return filas

    def _sincronizar(self, filas, prefix):
        from apps.integrantes.models import Integrante

        creados = 0
        actualizados = 0
        errores = []

        with transaction.atomic():
            for d in filas:
                nombre = d['nombre']
                nick   = d['nick']
                estado = d['estado']

                self.stdout.write(
                    f"  {prefix}{nombre} ({nick}) [{d['estado_raw']}] -> {estado}"
                )

                if self.dry_run:
                    continue

                try:
                    obj = None

                    # 1. Buscar por nick exacto (case-insensitive)
                    if nick:
                        obj = Integrante.objects.filter(nick__iexact=nick).first()

                    # 2. Buscar por nombre exacto si no hubo match por nick
                    if obj is None and nombre:
                        obj = Integrante.objects.filter(nombre__iexact=nombre).first()

                    if obj:
                        # Actualizar solo nombre, nick y estado — no tocar datos financieros ni contacto
                        obj.nombre = nombre
                        obj.nick   = nick
                        obj.estado = estado
                        obj.observaciones = (
                            obj.observaciones + '\n[Actualizado desde MENSUALIDADES 2025]'
                            if obj.observaciones and '[Actualizado desde MENSUALIDADES 2025]' not in obj.observaciones
                            else obj.observaciones or '[Actualizado desde MENSUALIDADES 2025]'
                        )
                        obj.save(update_fields=['nombre', 'nick', 'estado', 'observaciones', 'fecha_actualizacion'])
                        actualizados += 1
                    else:
                        Integrante.objects.create(
                            nombre=nombre,
                            nick=nick,
                            estado=estado,
                            observaciones='Sincronizado desde MENSUALIDADES 2025',
                        )
                        creados += 1

                except Exception as e:
                    errores.append(f'"{nombre}" ({nick}): {e}')

            if self.dry_run:
                transaction.set_rollback(True)

        if self.dry_run:
            self.stdout.write(self.style.SUCCESS(
                f'\n[DRY-RUN] Se procesarían {len(filas)} integrantes. '
                'No se guardó nada.'
            ))
            return

        self.stdout.write(self.style.SUCCESS(
            f'\nSincronizacion completa: {creados} creados, {actualizados} actualizados'
        ))

        if errores:
            self.stdout.write(self.style.ERROR(f'\nErrores ({len(errores)}):'))
            for e in errores:
                self.stdout.write(f'  - {e}')

        # Resumen por estado
        from apps.integrantes.models import Integrante
        self.stdout.write('\nEstado actual de la BD:')
        for code, label in Integrante.ESTADO_CHOICES:
            n = Integrante.objects.filter(estado=code).count()
            if n:
                self.stdout.write(f'  {label}: {n}')
        self.stdout.write(f'  Total: {Integrante.objects.count()}')
