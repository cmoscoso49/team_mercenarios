"""
Importa datos históricos desde el Excel del Team Mercenarios.
Uso: python manage.py importar_excel_historico
"""
import openpyxl
from datetime import datetime, date
from django.core.management.base import BaseCommand
from django.db import transaction


EXCEL_PATH = r"C:\Users\cmoscoso\OneDrive - INACAP\Descargas\2025\Mercenarios\datos team Actualizada 2022-2025 REVISADO POR ESTEBAN TL (4).xlsx"

MESES_MAP = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JUNIO ': 6, 'JULIO': 7,
    'AGOSTO': 8, 'AGOSTO ': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10,
    'NOBIEMBRE': 11, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
}

TALLA_MAP = {
    'XS': 'XS', 'S': 'S', 'M': 'M', 'L': 'L', 'XL': 'XL',
    'XXL': 'XXL', 'XXXL': 'XXL', '3XL': 'XXL',
}

ESTADO_MAP = {
    'ACTIVO': 'activo', 'INACTIVO': 'inactivo',
    'POS NATAL': 'honorario', 'POST NATAL': 'honorario',
    'SUSPENDIDO': 'suspendido', 'POSTULANTE': 'postulante',
}


def _val(v):
    """Normaliza valor de celda a string limpio."""
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ('None', 'nan') else s


def _rows_from(ws, start_row):
    """Itera filas con al menos 2 valores no vacíos."""
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        vals = [v for v in row if v is not None and str(v).strip()]
        if len(vals) >= 2:
            yield row


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s.split('.')[0], fmt).date()
        except ValueError:
            continue
    return None


def _normalize_nick(nick):
    return _val(nick).upper().strip() if nick else ''


class Command(BaseCommand):
    help = 'Importa datos históricos desde Excel. Elimina datos demo antes de importar.'

    def add_arguments(self, parser):
        parser.add_argument('--solo-mostrar', action='store_true',
                            help='Solo muestra lo que se importaría sin guardar')
        parser.add_argument('--sin-borrar-demo', action='store_true',
                            help='No elimina datos demo antes de importar')

    def handle(self, *args, **options):
        self.dry_run = options['solo_mostrar']
        self.errores = []
        self.stats = {}

        self.stdout.write(self.style.WARNING(
            f'{"[SIMULACIÓN] " if self.dry_run else ""}Iniciando importación histórica...\n'
        ))

        try:
            self.wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        except FileNotFoundError:
            self.stderr.write(f'ERROR: No se encontró el archivo:\n  {EXCEL_PATH}')
            return

        if not self.dry_run and not options['sin_borrar_demo']:
            self._limpiar_demo()

        with transaction.atomic():
            self._importar_integrantes()
            self._importar_mensualidades()
            self._importar_movimientos()
            self._importar_participaciones()
            self._importar_campeonatos()
            self._importar_reuniones()
            if self.dry_run:
                transaction.set_rollback(True)

        self.wb.close()
        self._imprimir_resumen()

    # ------------------------------------------------------------------ limpieza demo

    def _limpiar_demo(self):
        from apps.integrantes.models import Integrante
        from apps.finanzas.models import Movimiento, Mensualidad, Deuda
        from apps.eventos.models import Evento, Participacion
        from apps.noticias.models import Noticia
        from apps.finanzas.models import Categoria

        n_int = Integrante.objects.count()
        n_mov = Movimiento.objects.count()
        n_men = Mensualidad.objects.count()
        n_ev = Evento.objects.count()
        n_par = Participacion.objects.count()
        n_not = Noticia.objects.count()
        n_cat = Categoria.objects.count()

        Participacion.objects.all().delete()
        Mensualidad.objects.all().delete()
        Movimiento.objects.all().delete()
        Deuda.objects.all().delete()
        Evento.objects.all().delete()
        Noticia.objects.all().delete()
        Integrante.objects.all().delete()
        Categoria.objects.all().delete()

        self.stdout.write(f'  Demo eliminado: {n_int} integrantes, {n_mov} movimientos, '
                          f'{n_men} mensualidades, {n_ev} eventos, {n_par} participaciones, '
                          f'{n_not} noticias, {n_cat} categorías\n')

    # ------------------------------------------------------------------ integrantes

    def _importar_integrantes(self):
        from apps.integrantes.models import Integrante

        # Fuente principal: 'datos del team' (tiene fechas)
        # Enriquecimiento: 'DATOS + CEL' (tiene RUT + telefono)
        # Estado actual: 'DATOS ACTIVO'

        # 1. Construir índice nick→datos desde 'datos del team'
        datos_team = {}  # nick_upper → dict
        if 'datos del team' in self.wb.sheetnames:
            ws = self.wb['datos del team']
            # Headers: col4=nombre, col5=rut(vacío), col6=nick, col7=condicion, col8=talla, col9=cumpleaños, col10=ingreso
            for row in _rows_from(ws, 7):
                nombre = _val(row[3] if len(row) > 3 else None)
                nick = _val(row[5] if len(row) > 5 else None)
                if not nombre:
                    continue
                condicion = _val(row[6] if len(row) > 6 else None).upper()
                talla_raw = _val(row[7] if len(row) > 7 else None).upper()
                fecha_nac = _parse_date(row[8] if len(row) > 8 else None)
                fecha_ing = _parse_date(row[9] if len(row) > 9 else None)
                nick_key = _normalize_nick(nick) or nombre[:10].upper()
                datos_team[nick_key] = {
                    'nombre': nombre, 'nick': nick,
                    'estado': ESTADO_MAP.get(condicion, 'activo'),
                    'talla': TALLA_MAP.get(talla_raw, ''),
                    'fecha_nacimiento': fecha_nac,
                    'fecha_ingreso': fecha_ing,
                    'rut': '', 'telefono': '',
                }

        # 2. Enriquecer con RUT + teléfono desde 'DATOS + CEL'
        if 'DATOS + CEL' in self.wb.sheetnames:
            ws = self.wb['DATOS + CEL']
            # Headers R6: col4=nombre, col5=rut, col6=nick, col7=cancelado, col8=talla, col12=telefono
            for row in _rows_from(ws, 7):
                nombre = _val(row[3] if len(row) > 3 else None)
                rut = _val(row[4] if len(row) > 4 else None)
                nick = _val(row[5] if len(row) > 5 else None)
                telefono = _val(row[11] if len(row) > 11 else None)
                talla_raw = _val(row[7] if len(row) > 7 else None).upper()
                if not nombre:
                    continue
                nick_key = _normalize_nick(nick) or nombre[:10].upper()
                if nick_key in datos_team:
                    if rut:
                        datos_team[nick_key]['rut'] = rut
                    if telefono:
                        datos_team[nick_key]['telefono'] = telefono.replace(' ', '')
                    if talla_raw and not datos_team[nick_key]['talla']:
                        datos_team[nick_key]['talla'] = TALLA_MAP.get(talla_raw, '')
                else:
                    # Integrante nuevo solo en DATOS+CEL
                    datos_team[nick_key] = {
                        'nombre': nombre, 'nick': nick, 'rut': rut,
                        'telefono': telefono.replace(' ', ''),
                        'estado': 'activo',
                        'talla': TALLA_MAP.get(talla_raw, ''),
                        'fecha_nacimiento': None, 'fecha_ingreso': None,
                    }

        # 3. Actualizar estado actual desde 'DATOS ACTIVO'
        if 'DATOS ACTIVO' in self.wb.sheetnames:
            ws = self.wb['DATOS ACTIVO']
            for row in _rows_from(ws, 7):
                nick = _val(row[5] if len(row) > 5 else None)
                condicion = _val(row[6] if len(row) > 6 else None).upper()
                nick_key = _normalize_nick(nick)
                if nick_key in datos_team:
                    datos_team[nick_key]['estado'] = ESTADO_MAP.get(condicion, datos_team[nick_key]['estado'])

        # 4. Guardar
        creados = 0
        actualizados = 0
        for nick_key, d in datos_team.items():
            if not d['nombre']:
                continue
            rut = d['rut'] if d['rut'] else None
            talla = d['talla'] if d['talla'] in ('XS', 'S', 'M', 'L', 'XL', 'XXL') else ''
            defaults = {
                'nombre': d['nombre'][:150],
                'nick': (d['nick'] or '')[:50],
                'telefono': (d['telefono'] or '')[:20],
                'estado': d['estado'],
                'talla_polera': talla,
                'fecha_nacimiento': d['fecha_nacimiento'],
                'fecha_ingreso': d['fecha_ingreso'],
                'observaciones': 'Importado desde Excel histórico',
            }
            try:
                if rut:
                    obj, created = Integrante.objects.update_or_create(rut=rut, defaults=defaults)
                else:
                    # Sin RUT: buscar por nick
                    nick_norm = (d['nick'] or '')[:50]
                    obj, created = Integrante.objects.update_or_create(
                        nick=nick_norm, defaults=defaults
                    ) if nick_norm else (Integrante.objects.create(**defaults), True)
                if created:
                    creados += 1
                else:
                    actualizados += 1
            except Exception as e:
                self.errores.append(f'Integrante "{d["nombre"]}": {e}')

        self.stats['integrantes'] = {'creados': creados, 'actualizados': actualizados}
        self.stdout.write(f'  Integrantes: {creados} creados, {actualizados} actualizados')

    # ------------------------------------------------------------------ mensualidades

    def _importar_mensualidades(self):
        from apps.integrantes.models import Integrante
        from apps.finanzas.models import Mensualidad

        # Índice nick_upper → integrante
        nick_index = {i.nick.upper(): i for i in Integrante.objects.all() if i.nick}
        nombre_index = {i.nombre.upper(): i for i in Integrante.objects.all()}

        hojas_anios = {
            ' Mensaualidades 2022': 2022,
            'MENSUALIDADES 2023': 2023,
            'MENSUALIDADES 2024': 2024,
            'MENSUALIDADES 2025': 2025,
            'MENSUALIDADES 2026': 2026,
        }

        total_creadas = 0
        total_actualizadas = 0

        for hoja_nombre, anio in hojas_anios.items():
            if hoja_nombre not in self.wb.sheetnames:
                continue
            ws = self.wb[hoja_nombre]

            # Encontrar fila de cabecera (la que tiene 'NOMBRE DE PLAYERS' o 'NICK')
            header_row = None
            headers = []
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
                vals = [_val(v).upper() for v in row]
                if 'NICK' in vals or 'NOMBRE DE PLAYERS' in vals:
                    header_row = i
                    headers = vals
                    break

            if header_row is None:
                continue

            # Mapear columnas de meses
            mes_cols = {}  # mes_num → col_idx
            for ci, h in enumerate(headers):
                h_clean = h.strip().rstrip()
                mes = MESES_MAP.get(h_clean)
                if mes:
                    mes_cols[mes] = ci

            # Procesar filas de datos
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                nick_raw = _val(row[3] if len(row) > 3 else None) if len(row) > 3 else ''
                nombre_raw = _val(row[2] if len(row) > 2 else None) if len(row) > 2 else ''
                if not nick_raw and not nombre_raw:
                    continue
                # Parar si la fila es total/resumen
                if any(k in nick_raw.upper() for k in ('TOTAL', 'RESUMEN', 'SUMA')):
                    break

                # Buscar integrante
                integrante = nick_index.get(_normalize_nick(nick_raw))
                if not integrante and nombre_raw:
                    integrante = nombre_index.get(nombre_raw.upper())
                if not integrante:
                    # Buscar parcial por nick
                    for key, val in nick_index.items():
                        if nick_raw.upper() in key or key in nick_raw.upper():
                            integrante = val
                            break
                if not integrante:
                    self.errores.append(f'Mensualidad {anio}: nick "{nick_raw}" no encontrado')
                    continue

                for mes, col in mes_cols.items():
                    if col >= len(row):
                        continue
                    cell_val = _val(row[col])
                    if not cell_val:
                        estado = 'pendiente'
                    elif cell_val.replace('.', '').replace(',', '').isdigit():
                        estado = 'pagada'
                    elif 'GRACIA' in cell_val.upper() or 'EXENT' in cell_val.upper():
                        estado = 'exento'
                    else:
                        estado = 'pendiente'

                    try:
                        monto = int(float(cell_val.replace('.', '').replace(',', ''))) if estado == 'pagada' else 5000
                    except (ValueError, AttributeError):
                        monto = 5000

                    obj, created = Mensualidad.objects.update_or_create(
                        integrante=integrante, anio=anio, mes=mes,
                        defaults={'estado': estado, 'monto': monto}
                    )
                    if created:
                        total_creadas += 1
                    else:
                        total_actualizadas += 1

        self.stats['mensualidades'] = {'creadas': total_creadas, 'actualizadas': total_actualizadas}
        self.stdout.write(f'  Mensualidades: {total_creadas} creadas, {total_actualizadas} actualizadas')

    # ------------------------------------------------------------------ movimientos

    def _importar_movimientos(self):
        from apps.finanzas.models import Movimiento, Categoria

        # Crear categorías base
        cat_gastos, _ = Categoria.objects.get_or_create(nombre='Gastos Históricos', tipo='egreso')
        cat_ingresos, _ = Categoria.objects.get_or_create(nombre='Ingresos Históricos', tipo='ingreso')
        cat_caja, _ = Categoria.objects.get_or_create(nombre='Caja Chica', tipo='egreso')
        cat_donacion, _ = Categoria.objects.get_or_create(nombre='Donaciones', tipo='ingreso')
        cat_mensual, _ = Categoria.objects.get_or_create(nombre='Mensualidades', tipo='ingreso')

        creados = 0

        # --- H.CONTABLE ---
        if 'H.CONTABLE' in self.wb.sheetnames:
            ws = self.wb['H.CONTABLE']
            # Estructura: R7=headers (GASTOS: col2=N, col3=DETALLE, col4=FECHA, col5=VALOR UNIT, col6=CANTIDAD, col7=VALOR TOTAL)
            #                        (INGRESOS: col10=N, col11=DETALLE, col12=VALOR UNIT)
            for row in ws.iter_rows(min_row=8, values_only=True):
                if row is None or len(row) < 7:
                    continue

                # GASTOS (cols 1-6, índice base 0)
                detalle_g = _val(row[2] if len(row) > 2 else None)
                valor_g = _val(row[6] if len(row) > 6 else None)
                fecha_g = _parse_date(row[3] if len(row) > 3 else None)

                if detalle_g and valor_g:
                    try:
                        monto = abs(int(float(valor_g.replace('.', '').replace(',', ''))))
                        if monto > 0:
                            Movimiento.objects.create(
                                tipo='egreso', monto=monto,
                                descripcion=detalle_g[:200],
                                categoria=cat_gastos,
                                fecha=fecha_g or date(2022, 1, 1),
                                observaciones='Importado de H.CONTABLE',
                            )
                            creados += 1
                    except (ValueError, AttributeError):
                        pass

                # INGRESOS (cols 9-11)
                detalle_i = _val(row[10] if len(row) > 10 else None)
                valor_i = _val(row[11] if len(row) > 11 else None)

                if detalle_i and valor_i:
                    cat_ing = cat_donacion if 'DONAC' in detalle_i.upper() else (
                        cat_mensual if 'MENSUAL' in detalle_i.upper() else cat_ingresos
                    )
                    try:
                        monto = abs(int(float(valor_i.replace('.', '').replace(',', ''))))
                        if monto > 0:
                            Movimiento.objects.create(
                                tipo='ingreso', monto=monto,
                                descripcion=detalle_i[:200],
                                categoria=cat_ing,
                                fecha=date(2022, 1, 1),
                                observaciones='Importado de H.CONTABLE',
                            )
                            creados += 1
                    except (ValueError, AttributeError):
                        pass

        # --- GASTOS Y CAJA CHIK ---
        if 'GASTOS Y CAJA CHIK' in self.wb.sheetnames:
            ws = self.wb['GASTOS Y CAJA CHIK']
            for i, row in enumerate(ws.iter_rows(min_row=10, values_only=True), 10):
                if not row or len(row) < 5:
                    continue
                detalle = _val(row[3] if len(row) > 3 else None)
                monto_raw = _val(row[4] if len(row) > 4 else None)
                if not detalle or not monto_raw:
                    continue
                if any(k in detalle.upper() for k in ('TOTAL', 'CUADRATURA', 'MENSUAL')):
                    continue
                try:
                    monto = abs(int(float(monto_raw.replace('.', '').replace(',', ''))))
                    if monto > 0:
                        Movimiento.objects.create(
                            tipo='egreso', monto=monto,
                            descripcion=detalle[:200],
                            categoria=cat_caja,
                            fecha=date(2022, 1, 1),
                            es_caja_chica=True,
                            observaciones='Importado de GASTOS Y CAJA CHIK',
                        )
                        creados += 1
                except (ValueError, AttributeError):
                    pass

        self.stats['movimientos'] = {'creados': creados}
        self.stdout.write(f'  Movimientos: {creados} creados')

    # ------------------------------------------------------------------ participaciones

    def _importar_participaciones(self):
        from apps.integrantes.models import Integrante
        from apps.eventos.models import Evento, Participacion

        if 'LISTADO DE JUEGO' not in self.wb.sheetnames:
            self.stats['participaciones'] = {'creados': 0}
            return

        ws = self.wb['LISTADO DE JUEGO']
        nick_index = {i.nick.upper(): i for i in Integrante.objects.all() if i.nick}

        # R4: meses/períodos, R5: headers con fechas de juego, R7+: datos
        # Estructura: col1=N, col2=nombre, col3=nick, col4=estado, col5+=fechas
        # Necesitamos crear eventos y participaciones

        # Leer cabeceras de fechas (fila 5)
        header_row = None
        date_headers = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            vals = [_val(v) for v in row]
            if any('NICK' in v.upper() for v in vals if v):
                header_row = i
                date_headers = vals
                break

        if header_row is None:
            self.stats['participaciones'] = {'creados': 0}
            return

        # Crear evento genérico "Listado de Juego Histórico"
        evento, _ = Evento.objects.get_or_create(
            titulo='Partidas Históricas (2022-2025)',
            defaults={
                'tipo': 'partida',
                'fecha': date(2022, 1, 1),
                'estado': 'realizado',
                'descripcion': 'Participaciones históricas importadas desde Excel',
            }
        )

        creados = 0
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            if not row:
                continue
            nick_raw = _val(row[2] if len(row) > 2 else None)
            nombre_raw = _val(row[1] if len(row) > 1 else None)
            if not nick_raw and not nombre_raw:
                continue
            if any(k in str(nick_raw).upper() for k in ('TOTAL', 'SUMA', 'ACT/INACT')):
                break

            integrante = nick_index.get(_normalize_nick(nick_raw))
            if not integrante:
                continue

            # Contar partidas (celdas con 'X' o '1' o cualquier valor no vacío)
            total_partidas = sum(
                1 for v in row[4:] if _val(v) not in ('', 'ACT/INACT', 'INACTIVO', 'ACTIVO')
            )

            if total_partidas > 0:
                _, created = Participacion.objects.get_or_create(
                    integrante=integrante,
                    evento=evento,
                    defaults={'asistio': True, 'observaciones': f'Total partidas históricas: {total_partidas}'}
                )
                if created:
                    creados += 1

        self.stats['participaciones'] = {'creados': creados}
        self.stdout.write(f'  Participaciones: {creados} creadas')

    # ------------------------------------------------------------------ campeonatos

    def _importar_campeonatos(self):
        from apps.eventos.models import Evento, Participacion
        from apps.integrantes.models import Integrante

        if 'CAMPEONATO' not in self.wb.sheetnames:
            self.stats['campeonatos'] = {'creados': 0}
            return

        ws = self.wb['CAMPEONATO']
        nick_index = {i.nick.upper(): i for i in Integrante.objects.all() if i.nick}
        nombre_index = {}
        for i in Integrante.objects.all():
            # Índice por apellido+nombre parcial
            for part in i.nombre.upper().split():
                if len(part) > 4:
                    nombre_index[part] = i

        evento, _ = Evento.objects.get_or_create(
            titulo='Campeonato Vikingo 2022',
            defaults={
                'tipo': 'campeonato',
                'fecha': date(2022, 6, 1),
                'estado': 'realizado',
                'descripcion': 'Campeonato importado desde Excel histórico',
            }
        )

        creados = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            nombre_raw = _val(row[2] if len(row) > 2 else None)
            estado_raw = _val(row[4] if len(row) > 4 else None)
            if not nombre_raw:
                continue

            # Buscar integrante por nombre parcial
            integrante = None
            for part in nombre_raw.upper().split():
                if len(part) > 4:
                    integrante = nombre_index.get(part)
                    if integrante:
                        break

            if not integrante:
                continue

            asistio = 'OK' in estado_raw.upper() if estado_raw else False
            _, created = Participacion.objects.get_or_create(
                integrante=integrante,
                evento=evento,
                defaults={'asistio': asistio, 'observaciones': 'Campeonato Vikingo 2022'}
            )
            if created:
                creados += 1

        self.stats['campeonatos'] = {'creados': creados}
        self.stdout.write(f'  Campeonato (participaciones): {creados} creadas')

    # ------------------------------------------------------------------ reuniones

    def _importar_reuniones(self):
        from apps.eventos.models import Evento, Participacion
        from apps.integrantes.models import Integrante

        if 'reunion' not in self.wb.sheetnames:
            self.stats['reuniones'] = {'creados': 0}
            return

        ws = self.wb['reunion']
        nick_index = {i.nick.upper(): i for i in Integrante.objects.all() if i.nick}

        evento, _ = Evento.objects.get_or_create(
            titulo='Reunión Team Mercenarios (Histórica)',
            defaults={
                'tipo': 'reunion',
                'fecha': date(2022, 1, 1),
                'estado': 'realizado',
                'descripcion': 'Reunión importada desde Excel histórico',
            }
        )

        creados = 0
        for row in ws.iter_rows(min_row=8, values_only=True):
            nick_raw = _val(row[3] if len(row) > 3 else None)
            asistencia_raw = _val(row[5] if len(row) > 5 else None).upper()
            if not nick_raw:
                continue

            integrante = nick_index.get(_normalize_nick(nick_raw))
            if not integrante:
                continue

            asistio = asistencia_raw in ('SI', 'SÍ', 'S', 'ASISTIO', 'ASISTIÓ', 'YES')
            _, created = Participacion.objects.get_or_create(
                integrante=integrante,
                evento=evento,
                defaults={'asistio': asistio, 'observaciones': f'Reunión histórica: {asistencia_raw}'}
            )
            if created:
                creados += 1

        self.stats['reuniones'] = {'creados': creados}
        self.stdout.write(f'  Reunión (participaciones): {creados} creadas')

    # ------------------------------------------------------------------ resumen

    def _imprimir_resumen(self):
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=== RESUMEN DE MIGRACIÓN ==='))
        for modulo, datos in self.stats.items():
            parts = ', '.join(f'{k}: {v}' for k, v in datos.items())
            self.stdout.write(f'  {modulo.capitalize()}: {parts}')

        if self.errores:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(f'=== ERRORES ({len(self.errores)}) ==='))
            for e in self.errores[:30]:
                self.stdout.write(f'  - {e}')
            if len(self.errores) > 30:
                self.stdout.write(f'  ... y {len(self.errores) - 30} más')
        else:
            self.stdout.write(self.style.SUCCESS('  Sin errores'))

        self.stdout.write('')
        if self.dry_run:
            self.stdout.write(self.style.WARNING('[SIMULACIÓN] No se guardaron cambios.'))
        else:
            self.stdout.write(self.style.SUCCESS('Importación completada.'))
