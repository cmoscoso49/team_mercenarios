from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework.generics import ListAPIView
from .models import ImportacionArchivo
from .serializers import ImportacionArchivoSerializer


class ImportarExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        archivo = request.FILES.get('archivo')
        tipo = request.data.get('tipo', 'movimientos')

        if not archivo:
            return Response({'error': 'Se requiere un archivo'}, status=status.HTTP_400_BAD_REQUEST)

        nombre = archivo.name.lower()
        if not (nombre.endswith('.xlsx') or nombre.endswith('.xls') or nombre.endswith('.csv')):
            return Response({'error': 'Solo se aceptan archivos .xlsx, .xls o .csv'},
                            status=status.HTTP_400_BAD_REQUEST)

        importacion = ImportacionArchivo.objects.create(
            archivo=archivo,
            tipo=tipo,
            importado_por=request.user,
        )

        try:
            procesados, errores = self._procesar_archivo(importacion, tipo)
            importacion.estado = 'procesado'
            importacion.registros_procesados = procesados
            importacion.registros_con_error = errores
            importacion.mensaje = f"Procesados: {procesados}, Errores: {errores}"
            importacion.save()
            return Response({
                'mensaje': importacion.mensaje,
                'id': importacion.id,
                'registros_procesados': procesados,
                'registros_con_error': errores,
            })
        except Exception as e:
            importacion.estado = 'error'
            importacion.mensaje = str(e)
            importacion.save()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _procesar_archivo(self, importacion, tipo):
        import openpyxl
        from datetime import datetime

        nombre = importacion.archivo.name.lower()
        procesados = 0
        errores = 0

        if nombre.endswith('.csv'):
            import csv
            importacion.archivo.open('r')
            reader = csv.DictReader(importacion.archivo)
            rows = list(reader)
            importacion.archivo.close()
        else:
            wb = openpyxl.load_workbook(importacion.archivo)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

        for row in rows:
            try:
                if tipo == 'movimientos':
                    self._importar_movimiento(row)
                elif tipo == 'integrantes':
                    self._importar_integrante(row)
                elif tipo == 'mensualidades':
                    self._importar_mensualidad(row)
                procesados += 1
            except Exception:
                errores += 1

        return procesados, errores

    def _importar_movimiento(self, row):
        from .models import Movimiento, Categoria
        from datetime import datetime

        tipo = str(row.get('tipo', row.get('Tipo', 'egreso'))).lower()
        if 'ingreso' in tipo:
            tipo = 'ingreso'
        else:
            tipo = 'egreso'

        monto_raw = row.get('monto', row.get('Monto', 0))
        monto = abs(float(str(monto_raw).replace('$', '').replace('.', '').replace(',', '.') or 0))

        descripcion = str(row.get('descripcion', row.get('Descripcion', row.get('glosa', row.get('Glosa', 'Importado')))))
        fecha_raw = row.get('fecha', row.get('Fecha'))
        if isinstance(fecha_raw, str) and fecha_raw:
            try:
                fecha = datetime.strptime(fecha_raw, '%d/%m/%Y').date()
            except ValueError:
                try:
                    fecha = datetime.strptime(fecha_raw, '%Y-%m-%d').date()
                except ValueError:
                    fecha = datetime.today().date()
        elif hasattr(fecha_raw, 'date'):
            fecha = fecha_raw.date()
        else:
            fecha = datetime.today().date()

        Movimiento.objects.create(
            tipo=tipo,
            monto=monto,
            descripcion=descripcion[:200],
            fecha=fecha,
            observaciones='Importado desde archivo',
        )

    def _importar_integrante(self, row):
        from apps.integrantes.models import Integrante
        nombre = str(row.get('nombre', row.get('Nombre', ''))).strip()
        if not nombre:
            raise ValueError('Nombre requerido')

        Integrante.objects.get_or_create(
            nombre=nombre,
            defaults={
                'nick': str(row.get('nick', row.get('Nick', ''))),
                'telefono': str(row.get('telefono', row.get('Telefono', ''))),
                'email': str(row.get('email', row.get('Email', ''))),
                'observaciones': 'Importado desde archivo',
            }
        )

    def _importar_mensualidad(self, row):
        from apps.integrantes.models import Integrante
        from .models import Mensualidad
        nombre = str(row.get('nombre', row.get('Nombre', ''))).strip()
        integrante = Integrante.objects.filter(nombre__icontains=nombre).first()
        if not integrante:
            raise ValueError(f'Integrante no encontrado: {nombre}')

        anio = int(row.get('anio', row.get('Año', row.get('año', 2024))))
        mes = int(row.get('mes', row.get('Mes', 1)))
        estado = 'pagada' if str(row.get('estado', row.get('Estado', 'pendiente'))).lower() in ['pagado', 'pagada', 'p', 'si', 'sí', '1'] else 'pendiente'

        Mensualidad.objects.update_or_create(
            integrante=integrante, anio=anio, mes=mes,
            defaults={'estado': estado}
        )


class ImportacionListView(ListAPIView):
    queryset = ImportacionArchivo.objects.all()
    serializer_class = ImportacionArchivoSerializer
    permission_classes = [IsAuthenticated]
